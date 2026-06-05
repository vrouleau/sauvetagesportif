"""Inspect the SERC XLSX to understand structure."""
import openpyxl
from pathlib import Path

xlsx_path = Path(r"C:\Users\eoivnru\Documents\MeetManager\sauvetagesportif\serc_claude\SERC Generator CPLC 2026 version 28 mai 2026.xlsm")
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

print(f"Sheets: {wb.sheetnames}")
print()

# Factor Sheet
ws = wb["Factor Sheet"]
print("=== FACTOR SHEET ===")
for row in ws.iter_rows(min_row=1, max_row=30, max_col=20, values_only=False):
    cells = [(c.coordinate, c.value) for c in row if c.value is not None]
    if cells:
        print("  ", cells)
print()

# Team Entry
ws = wb["Team Entry"]
print("\n=== TEAM ENTRY ===")
for row in ws.iter_rows(min_row=1, max_row=35, max_col=20, values_only=False):
    cells = [(c.coordinate, c.value) for c in row if c.value is not None]
    if cells:
        print("  ", cells)
print()

# Scoring sheet
ws = wb["Scoring"]
print("\n=== SCORING (first 15 rows) ===")
for row in ws.iter_rows(min_row=1, max_row=15, max_col=15, values_only=False):
    cells = [(c.coordinate, c.value) for c in row if c.value is not None]
    if cells:
        print("  ", cells)
print()

# VICTIM_1 sheet
ws = wb["VICTIM_1"]
print("\n=== VICTIM_1 (first 20 rows) ===")
for row in ws.iter_rows(min_row=1, max_row=20, max_col=15, values_only=False):
    cells = [(c.coordinate, c.value) for c in row if c.value is not None]
    if cells:
        print("  ", cells)
print()

# OVERALL sheet
ws = wb["OVERALL"]
print("\n=== OVERALL (first 20 rows) ===")
for row in ws.iter_rows(min_row=1, max_row=20, max_col=15, values_only=False):
    cells = [(c.coordinate, c.value) for c in row if c.value is not None]
    if cells:
        print("  ", cells)

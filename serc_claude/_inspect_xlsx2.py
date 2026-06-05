"""Full inspection of all SERC XLSX sheets."""
import openpyxl
from pathlib import Path

xlsx_path = Path(r"C:\Users\eoivnru\Documents\MeetManager\sauvetagesportif\serc_claude\SERC Generator CPLC 2026 version 28 mai 2026.xlsm")
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

print(f"Sheets: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*70}")
    print(f"=== {sheet_name} ===")
    print(f"{'='*70}")
    max_row = min(ws.max_row or 50, 50)
    max_col = min(ws.max_column or 20, 20)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=False):
        cells = [(c.coordinate, c.value) for c in row if c.value is not None]
        if cells:
            print("  ", cells)
